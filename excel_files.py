import pandas as pd

from openpyxl import load_workbook   # для работы с файлами Excel (загрузка)
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill  # для работы с файлами Excel (форматирование)


def save_to_excel(result_df, no_matches_df, result_file_path):
    """сохранение результата в виде отформатированной Excel-таблицы"""

    with pd.ExcelWriter(result_file_path) as writer:
        result_df.to_excel(writer, sheet_name='Основная таблица', startrow=1, index=False)
        no_matches_df.to_excel(writer, sheet_name='Основная таблица', startrow=len(result_df) + 4, startcol=6, index=True)

    wb = load_workbook(result_file_path)
    ws = wb.active

    format_excel(ws, result_df, no_matches_df)
    wb.save(result_file_path)


def format_excel(ws, result_df, no_matches_df):
    """Форматирование Excel-таблицы"""
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    bold_font = Font(bold=True)    # Шрифт (жирный)

    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),   # Границы ячеек
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    # Применяем выравнивание и границы для заголовков
    for cell in ws[1]:
        cell.alignment = alignment
        cell.font = bold_font
        cell.border = border

    # Применяем выравнивание, границы и перенос строки для всех остальных ячеек
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # Удаляем границы ячеек между таблицами, чтобы сделать визуальный отступ
    start_row = len(result_df) + 3
    end_row = start_row + len(no_matches_df) + 6

    for row in range(start_row, start_row + 3):
        for cell in ws[row]:
            cell.border = Border()

    # Удаляем границы столбцов с пустыми ячейками во второй таблице (no_matches_df)
    for row in range(start_row, end_row):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'M', 'N', 'O', 'P', 'Q', 'R']:  # список столбцов без границ
            cell = ws[f'{col}{row}']
            cell.border = Border()

    # Применяем окрашивание строк при наличии соответствующих условий:
    # зеленый - строка установлена правильно,
    # желтый - строка без совпадения,
    # красный - строка имеет ошибку в совпадении, менеджеру нужно обратить на это внимание

    yellow_fill = PatternFill(start_color="FFFEA8", end_color="FFFEA8", fill_type="solid")  # светло-желтый
    red_fill = PatternFill(start_color="FFA8A8", end_color="FFA8A8", fill_type="solid")  # Светло-красный
    green_fill = PatternFill(start_color="A8FFBF", end_color="A8FFBF", fill_type="solid")  # Светло-красный

    # Проходим по строкам и проверяем условие
    for row in range(3, len(result_df) + 3):
        name = ws.cell(row=row, column=8).value  # Столбец 'Наименование номенклатуры в счете' (восьмой столбец)
        price = ws.cell(row=row, column=16).value  # Столбец "Несоответствие в цене" (16 столбец)
        quantity = ws.cell(row=row, column=17).value  # Столбец "Несоответствие в количестве" (17 столбец)
        unit = ws.cell(row=row, column=18).value  # Столбец "Несоответствие в ед.изм." (18 столбец)

        # Если в 'Наименование номенклатуры в заявке' есть значение, а в 'Наименование номенклатуры в счете' пусто,
        # Окрашиваем строку в желтый цвет, пропуская столбцы 7 и 13 (являются разграничителями)
        if not name:
            for col in range(1, ws.max_column + 1):
                if col not in [7, 13]:
                    ws.cell(row=row, column=col).fill = yellow_fill

        # Если в столбце "Несоответствие в цене/количестве/ед.изм" есть значение,
        # Окрашиваем строку в красный цвет
        elif price or quantity or unit:
            for col in range(1, ws.max_column + 1):
                if col not in [7, 13]:
                    ws.cell(row=row, column=col).fill = red_fill

        # Все оставшиеся строки окрашиваем в зеленый цвет
        else:
            for col in range(1, ws.max_column + 1):
                if col not in [7, 13]:
                    ws.cell(row=row, column=col).fill = green_fill

    # Применяем заливку и убираем границы для столбцов 7 и 13 (для визуального деления таблицы)
    for row in range(1, len(result_df) + 3):  # Начинаем с 2 строки, если есть заголовки
        for col in [7, 13]:
            cell = ws.cell(row=row, column=col)
            cell.border = Border()

    # Очищаем текст в первой ячейке столбца 7 и 13
    ws.cell(row=2, column=7).value = None
    ws.cell(row=2, column=13).value = None

    ws.merge_cells('A1:F1')  # Объединяем ячейки для "Заявка"
    ws.merge_cells('H1:L1')  # Объединяем ячейки для "Счет"
    ws.merge_cells('N1:R1')  # Объединяем ячейки для "Несоответствия"

    # Добавляем текст для этих объединённых ячеек
    ws['A1'] = 'Заявка'
    ws['H1'] = 'Счет'
    ws['N1'] = 'Несоответствие'

    # Центрируем текст в объединённых ячейках
    for cell in ['A1', 'H1', 'N1']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # Настраиваем ширину столбцов
    column_widths = {
        '№': 5,
        ' Наименование номенклатуры ': 50,
        ' Цена ': 10,
        ' Количество ': 12,
        ' Ед.изм. ': 10,
        ' Поставщик ': 20,
        '1': 3,
        'Наименование номенклатуры': 50,
        'Цена': 10,
        'Количество': 12,
        'Ед.изм.': 10,
        'Поставщик': 20,
        '2': 3,
        'Дубли и схожие позиции': 30,
        'в тексте': 15,
        'в цене': 15,
        'в количестве': 15,
        'в ед.изм.': 15
    }

    for idx, col_name in enumerate(result_df.columns, start=1):
        if idx > 1:  # Пропускаем первую строку с объединенными ячейками
            col_letter = chr(64 + idx)  # Преобразуем индекс столбца в букву
            ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)  # Устанавливаем ширину
